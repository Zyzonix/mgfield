#!/usr/bin/python3
#
# written by @author ZyzonixDev
# published by OCISLY and ZyzonixDevelopments
# -
# date      | 25/04/2021
# python-v  | 3.5.3
# -
# file      | static/xlsxHandler.py
# project   | MGFieldPy
# project-v | 1.1
# 
from datetime import date, datetime
import openpyxl

# Vorbereiten der Excel-Datei
def prepareXLSXFile(filePath, sheetName, fileName):
    row = 1
    # Initialisiert die notwendige Bibliothek zum schreiben/lesen von Excel-Dateien
    workBook = openpyxl.Workbook()
    # Aktiviert die zu bearbeitende Tabelle
    workSheet = workBook.active
    # Setzt den Tabellennamen
    workSheet.title = sheetName
    # Variable für die Spaltenbezeichnung (Zeile 1)
    values = ["Date", "Time", "MGField Value1","MGField Value2", "Temperature", "Ram-Available"]
    # Schreibt die vorher behannten Bezeichnungen in die erste Zeile
    for column, item in enumerate(values, start=1):
        workSheet.cell(row, column, item)
    # Speichert die Datei
    workBook.save(filename=filePath)
    # Dateiname als self.-Variable gesichert
    return fileName

# Schreiben von Daten in die Datei
def writeToXLSXFile(self, data):
    # Öffnen der Datei und aktivieren der Tabelle
    workBook = openpyxl.load_workbook(self.fileData[2])
    workSheet = workBook[self.fileData[1]]
    currentRow = workBook.active.max_row + 1
    # Schreibt Daten in die Zeile
    for column, item in enumerate(data, start=1):
        workSheet.cell(currentRow, column, item)
    # Speichert die Datei
    workBook.save(filename=self.fileData[2])
